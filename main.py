from __future__ import annotations

import csv
import os
import re
import sqlite3
import sys
import time
import loggingtweewttew
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Iterable, Optional, Sequence

import yaml
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# ----------------------------
# Logging
# ----------------------------
LOG = logging.getLogger("yt_watch")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
)


# ----------------------------
# Models
# ----------------------------
@dataclass(frozen=True)
class Config:
    api_key_env: str
    scan_limit_per_channel: int
    take_last_videos: int
    take_last_shorts: int
    shorts_max_seconds: int
    require_hashtag_shorts: bool
    refresh_stats_days: int
    output_xlsx: str
    timezone: str


@dataclass(frozen=True)
class VideoRow:
    channel_id: str
    channel_title: str
    video_id: str
    title: str
    published_at: datetime
    duration_sec: int
    is_shorts: bool
    url: str
    view_count: Optional[int]
    like_count: Optional[int]
    comment_count: Optional[int]


# ----------------------------
# Utils
# ----------------------------
def utcnow() -> datetime:
    return datetime.now(timezone.utc)


def parse_rfc3339(dt_str: str) -> datetime:
    # example: 2025-01-01T12:34:56Z
    return datetime.fromisoformat(dt_str.replace("Z", "+00:00")).astimezone(timezone.utc)


def parse_iso8601_duration_to_seconds(s: str) -> int:
    # PT#H#M#S
    m = re.fullmatch(r"PT(?:(\d+)H)?(?:(\d+)M)?(?:(\d+)S)?", s or "")
    if not m:
        return 0
    h = int(m.group(1) or 0)
    mi = int(m.group(2) or 0)
    sec = int(m.group(3) or 0)
    return h * 3600 + mi * 60 + sec


def normalize_channel_source(raw: str) -> str:
    s = (raw or "").strip()
    return s.rstrip("/")


def extract_channel_id_or_handle(raw: str) -> tuple[Optional[str], Optional[str]]:
    """
    Returns (channel_id, handle).
    Supports:
      - UC...
      - @handle
      - https://www.youtube.com/@handle
      - https://www.youtube.com/channel/UC...
    """
    s = normalize_channel_source(raw)

    if s.startswith("UC") and len(s) >= 20:
        return s, None

    if s.startswith("@") and len(s) > 1:
        return None, s[1:]

    m = re.search(r"youtube\.com/@([^/?#]+)", s)
    if m:
        return None, m.group(1)

    m = re.search(r"youtube\.com/channel/(UC[^/?#]+)", s)
    if m:
        return m.group(1), None

    # Не “угадываем” /c/... потому что это может потребовать дорогого поиска.
    return None, None


def is_probably_shorts(
    *,
    duration_sec: int,
    shorts_max_seconds: int,
    require_hashtag_shorts: bool,
    title: str,
    description: str,
    tags: Sequence[str],
) -> bool:
    if duration_sec <= 0 or duration_sec > shorts_max_seconds:
        return False

    if not require_hashtag_shorts:
        return True

    blob = f"{title}\n{description}".lower()
    if "#shorts" in blob:
        return True

    for t in tags:
        tt = (t or "").lower().strip()
        if tt in ("shorts", "#shorts"):
            return True

    return False


# ----------------------------
# DB
# ----------------------------
class DB:
    def __init__(self, path: Path):
        self.path = path
        self.con = sqlite3.connect(str(path))
        self.con.row_factory = sqlite3.Row
        self._migrate()

    def close(self) -> None:
        self.con.close()

    def _migrate(self) -> None:
        cur = self.con.cursor()
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS channels (
                channel_id TEXT PRIMARY KEY,
                source TEXT NOT NULL,
                title TEXT,
                uploads_playlist_id TEXT,
                last_sync_at TEXT
            )
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS videos (
                video_id TEXT PRIMARY KEY,
                channel_id TEXT NOT NULL,
                title TEXT,
                published_at TEXT,
                duration_sec INTEGER,
                is_shorts INTEGER,
                url TEXT,
                description TEXT,
                tags_json TEXT,
                view_count INTEGER,
                like_count INTEGER,
                comment_count INTEGER,
                stats_updated_at TEXT,
                FOREIGN KEY(channel_id) REFERENCES channels(channel_id)
            )
            """
        )
        cur.execute("CREATE INDEX IF NOT EXISTS idx_videos_channel_pub ON videos(channel_id, published_at)")
        self.con.commit()

    def upsert_channel(self, channel_id: str, source: str, title: str, uploads_playlist_id: str) -> None:
        self.con.execute(
            """
            INSERT INTO channels(channel_id, source, title, uploads_playlist_id, last_sync_at)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(channel_id) DO UPDATE SET
                source=excluded.source,
                title=excluded.title,
                uploads_playlist_id=excluded.uploads_playlist_id
            """,
            (channel_id, source, title, uploads_playlist_id, None),
        )
        self.con.commit()

    def set_channel_synced(self, channel_id: str) -> None:
        self.con.execute(
            "UPDATE channels SET last_sync_at=? WHERE channel_id=?",
            (utcnow().isoformat(), channel_id),
        )
        self.con.commit()

    def get_known_video_ids(self, channel_id: str) -> set[str]:
        cur = self.con.execute("SELECT video_id FROM videos WHERE channel_id=?", (channel_id,))
        return {row["video_id"] for row in cur.fetchall()}

    def upsert_video(self, row: VideoRow, description: str, tags: Sequence[str]) -> None:
        # tags store simple json-ish string without bringing extra deps
        tags_json = "[" + ",".join([sqlite3.escape_string((t or "").encode("utf-8")).decode("utf-8") for t in tags]) + "]"
        self.con.execute(
            """
            INSERT INTO videos(
                video_id, channel_id, title, published_at, duration_sec, is_shorts, url,
                description, tags_json,
                view_count, like_count, comment_count, stats_updated_at
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(video_id) DO UPDATE SET
                channel_id=excluded.channel_id,
                title=excluded.title,
                published_at=excluded.published_at,
                duration_sec=excluded.duration_sec,
                is_shorts=excluded.is_shorts,
                url=excluded.url,
                description=excluded.description,
                tags_json=excluded.tags_json,
                view_count=excluded.view_count,
                like_count=excluded.like_count,
                comment_count=excluded.comment_count,
                stats_updated_at=excluded.stats_updated_at
            """,
            (
                row.video_id,
                row.channel_id,
                row.title,
                row.published_at.isoformat(),
                row.duration_sec,
                1 if row.is_shorts else 0,
                row.url,
                description or "",
                tags_json,
                row.view_count,
                row.like_count,
                row.comment_count,
                utcnow().isoformat(),
            ),
        )
        self.con.commit()

    def needs_stats_refresh(self, video_id: str, max_age_days: int) -> bool:
        cur = self.con.execute("SELECT stats_updated_at FROM videos WHERE video_id=?", (video_id,))
        row = cur.fetchone()
        if not row:
            return True
        s = row["stats_updated_at"]
        if not s:
            return True
        try:
            dt = datetime.fromisoformat(s)
        except ValueError:
            return True
        return (utcnow() - dt.replace(tzinfo=timezone.utc)) > timedelta(days=max_age_days)

    def select_export_rows(self) -> tuple[list[sqlite3.Row], list[sqlite3.Row], list[sqlite3.Row]]:
        # All, Videos (not shorts), Shorts
        all_rows = self.con.execute(
            """
            SELECT v.*, c.title as channel_title
            FROM videos v
            JOIN channels c ON c.channel_id = v.channel_id
            ORDER BY datetime(v.published_at) DESC
            """
        ).fetchall()

        videos = self.con.execute(
            """
            SELECT v.*, c.title as channel_title
            FROM videos v
            JOIN channels c ON c.channel_id = v.channel_id
            WHERE v.is_shorts=0
            ORDER BY c.title, datetime(v.published_at) DESC
            """
        ).fetchall()

        shorts = self.con.execute(
            """
            SELECT v.*, c.title as channel_title
            FROM videos v
            JOIN channels c ON c.channel_id = v.channel_id
            WHERE v.is_shorts=1
            ORDER BY c.title, datetime(v.published_at) DESC
            """
        ).fetchall()

        return all_rows, videos, shorts


# ----------------------------
# YouTube Client (quota-friendly)
# ----------------------------
class YouTubeClient:
    def __init__(self, api_key: str):
        self.yt = build("youtube", "v3", developerKey=api_key, cache_discovery=False)

    def execute(self, req, *, retries: int = 3):
        # googleapiclient умеет num_retries
        return req.execute(num_retries=retries)

    def resolve_channel(self, source: str) -> tuple[str, str, str]:
        """
        Returns (channel_id, channel_title, uploads_playlist_id).
        Uses channels.list (cost 1 unit). Default quota: 10k/day.
        """
        channel_id, handle = extract_channel_id_or_handle(source)

        if channel_id:
            resp = self.execute(
                self.yt.channels().list(
                    part="snippet,contentDetails",
                    id=channel_id,
                    maxResults=1,
                    fields="items(id,snippet(title),contentDetails(relatedPlaylists(uploads)))",
                )
            )
        elif handle:
            resp = self.execute(
                self.yt.channels().list(
                    part="snippet,contentDetails",
                    forHandle=handle,
                    maxResults=1,
                    fields="items(id,snippet(title),contentDetails(relatedPlaylists(uploads)))",
                )
            )
        else:
            raise ValueError(
                f"Не смог распознать канал: {source}\n"
                f"Используй UC... / @handle / ссылку вида youtube.com/@handle или youtube.com/channel/UC..."
            )

        items = resp.get("items", [])
        if not items:
            raise ValueError(f"Канал не найден: {source}")

        it = items[0]
        cid = it["id"]
        title = it["snippet"]["title"]
        uploads = it["contentDetails"]["relatedPlaylists"]["uploads"]
        return cid, title, uploads

    def fetch_upload_video_ids(self, uploads_playlist_id: str, limit: int) -> list[str]:
        out: list[str] = []
        page_token = None

        while len(out) < limit:
            req = self.yt.playlistItems().list(
                part="contentDetails",
                playlistId=uploads_playlist_id,
                maxResults=min(50, limit - len(out)),
                pageToken=page_token,
                fields="items(contentDetails(videoId)),nextPageToken",
            )
            resp = self.execute(req)
            for it in resp.get("items", []):
                out.append(it["contentDetails"]["videoId"])

            page_token = resp.get("nextPageToken")
            if not page_token:
                break

        return out

    def fetch_videos_details(self, video_ids: Sequence[str]) -> list[dict]:
        """
        Returns raw video resources.
        videos.list cost = 1 unit. Avoid search.list (100 units).
        """
        if not video_ids:
            return []

        results: list[dict] = []
        for i in range(0, len(video_ids), 50):
            chunk = video_ids[i : i + 50]
            req = self.yt.videos().list(
                part="snippet,contentDetails,statistics",
                id=",".join(chunk),
                fields=(
                    "items(id,"
                    "snippet(title,publishedAt,description,tags),"
                    "contentDetails(duration),"
                    "statistics(viewCount,likeCount,commentCount))"
                ),
            )
            resp = self.execute(req)
            results.extend(resp.get("items", []))

        return results


# ----------------------------
# Excel export
# ----------------------------
def export_xlsx(
    *,
    db: DB,
    out_path: Path,
    take_last_videos: int,
    take_last_shorts: int,
) -> None:
    wb = Workbook()
    ws_v = wb.active
    ws_v.title = "Latest Videos"
    ws_s = wb.create_sheet("Latest Shorts")
    ws_all = wb.create_sheet("All (desc)")

    headers = [
        "Channel",
        "Channel ID",
        "Type",
        "Title",
        "Published (UTC)",
        "Duration (sec)",
        "Views",
        "Likes",
        "Comments",
        "URL",
    ]

    def style_sheet(ws):
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            ws.cell(row=1, column=c).font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    style_sheet(ws_v)
    style_sheet(ws_s)
    style_sheet(ws_all)

    all_rows, videos_rows, shorts_rows = db.select_export_rows()

    # Latest 5+5 per channel
    latest_v: dict[str, int] = {}
    latest_s: dict[str, int] = {}

    for r in videos_rows:
        cid = r["channel_id"]
        if latest_v.get(cid, 0) >= take_last_videos:
            continue
        latest_v[cid] = latest_v.get(cid, 0) + 1
        ws_v.append(
            [
                r["channel_title"],
                cid,
                "video",
                r["title"],
                r["published_at"],
                r["duration_sec"],
                r["view_count"],
                r["like_count"],
                r["comment_count"],
                r["url"],
            ]
        )

    for r in shorts_rows:
        cid = r["channel_id"]
        if latest_s.get(cid, 0) >= take_last_shorts:
            continue
        latest_s[cid] = latest_s.get(cid, 0) + 1
        ws_s.append(
            [
                r["channel_title"],
                cid,
                "shorts",
                r["title"],
                r["published_at"],
                r["duration_sec"],
                r["view_count"],
                r["like_count"],
                r["comment_count"],
                r["url"],
            ]
        )

    for r in all_rows:
        ws_all.append(
            [
                r["channel_title"],
                r["channel_id"],
                "shorts" if r["is_shorts"] else "video",
                r["title"],
                r["published_at"],
                r["duration_sec"],
                r["view_count"],
                r["like_count"],
                r["comment_count"],
                r["url"],
            ]
        )

    def autosize(ws):
        for col in range(1, len(headers) + 1):
            max_len = 0
            letter = get_column_letter(col)
            for cell in ws[letter]:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[letter].width = min(max_len + 2, 70)

    autosize(ws_v)
    autosize(ws_s)
    autosize(ws_all)

    wb.save(out_path)


# ----------------------------
# IO
# ----------------------------
def load_config(path: Path) -> Config:
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    return Config(
        api_key_env=str(data.get("api_key_env", "YOUTUBE_API_KEY")),
        scan_limit_per_channel=int(data.get("scan_limit_per_channel", 200)),
        take_last_videos=int(data.get("take_last_videos", 5)),
        take_last_shorts=int(data.get("take_last_shorts", 5)),
        shorts_max_seconds=int(data.get("shorts_max_seconds", 180)),
        require_hashtag_shorts=bool(data.get("require_hashtag_shorts", False)),
        refresh_stats_days=int(data.get("refresh_stats_days", 30)),
        output_xlsx=str(data.get("output_xlsx", "monitor.xlsx")),
        timezone=str(data.get("timezone", "UTC")),
    )


def load_channels_csv(path: Path) -> list[str]:
    rows: list[str] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if "channel" not in reader.fieldnames:
            raise ValueError("channels.csv должен иметь колонку 'channel'")
        for r in reader:
            s = normalize_channel_source(r.get("channel", ""))
            if s:
                rows.append(s)
    # unique but stable order
    seen = set()
    uniq: list[str] = []
    for x in rows:
        if x not in seen:
            seen.add(x)
            uniq.append(x)
    return uniq


# ----------------------------
# Main sync
# ----------------------------
def sync_channel(
    *,
    cfg: Config,
    db: DB,
    yt: YouTubeClient,
    source: str,
) -> None:
    channel_id, channel_title, uploads_pl = yt.resolve_channel(source)
    db.upsert_channel(channel_id, source, channel_title, uploads_pl)

    known = db.get_known_video_ids(channel_id)
    ids = yt.fetch_upload_video_ids(uploads_pl, limit=cfg.scan_limit_per_channel)

    # берём только новые + те, кому надо обновить статистику
    to_fetch: list[str] = []
    for vid in ids:
        if vid not in known:
            to_fetch.append(vid)
        else:
            # обновляем stats только если устарели (чтобы не “жечь” вызовы)
            if db.needs_stats_refresh(vid, max_age_days=cfg.refresh_stats_days):
                to_fetch.append(vid)

    if not to_fetch:
        LOG.info("OK (no changes): %s", channel_title)
        db.set_channel_synced(channel_id)
        return

    raw_videos = yt.fetch_videos_details(to_fetch)

    # Сохраняем
    for v in raw_videos:
        sn = v.get("snippet", {})
        cd = v.get("contentDetails", {})
        st = v.get("statistics", {})

        duration_sec = parse_iso8601_duration_to_seconds(cd.get("duration", ""))
        title = sn.get("title", "") or ""
        desc = sn.get("description", "") or ""
        tags = sn.get("tags", []) or []
        pub = parse_rfc3339(sn.get("publishedAt", "1970-01-01T00:00:00Z"))

        is_shorts = is_probably_shorts(
            duration_sec=duration_sec,
            shorts_max_seconds=cfg.shorts_max_seconds,
            require_hashtag_shorts=cfg.require_hashtag_shorts,
            title=title,
            description=desc,
            tags=tags,
        )

        row = VideoRow(
            channel_id=channel_id,
            channel_title=channel_title,
            video_id=v["id"],
            title=title,
            published_at=pub,
            duration_sec=duration_sec,
            is_shorts=is_shorts,
            url=f"https://www.youtube.com/watch?v={v['id']}",
            view_count=int(st["viewCount"]) if "viewCount" in st else None,
            like_count=int(st["likeCount"]) if "likeCount" in st else None,
            comment_count=int(st["commentCount"]) if "commentCount" in st else None,
        )
        db.upsert_video(row, description=desc, tags=tags)

    db.set_channel_synced(channel_id)
    LOG.info("OK: %s | fetched=%d", channel_title, len(raw_videos))


def main() -> int:
    base = Path(".").resolve()

    cfg_path = base / "config.yaml"
    channels_path = base / "channels.csv"
    db_path = base / "yt_watch.sqlite3"

    if not cfg_path.exists() or not channels_path.exists():
        LOG.error("Нужны файлы рядом со скриптом: config.yaml и channels.csv")
        return 2

    cfg = load_config(cfg_path)

    api_key = os.getenv(cfg.api_key_env)
    if not api_key:
        LOG.error("Не найден API key. Установи переменную окружения %s", cfg.api_key_env)
        return 2

    channels = load_channels_csv(channels_path)
    if not channels:
        LOG.error("channels.csv пустой")
        return 2

    yt = YouTubeClient(api_key=api_key)
    db = DB(db_path)

    try:
        for idx, source in enumerate(channels, 1):
            try:
                sync_channel(cfg=cfg, db=db, yt=yt, source=source)
                # маленькая пауза для “вежливости” к API
                time.sleep(0.05)
            except (HttpError, ValueError) as e:
                LOG.error("FAIL [%d/%d] %s | %s", idx, len(channels), source, e)

        out_xlsx = base / cfg.output_xlsx
        export_xlsx(
            db=db,
            out_path=out_xlsx,
            take_last_videos=cfg.take_last_videos,
            take_last_shorts=cfg.take_last_shorts,
        )
        LOG.info("Saved: %s", out_xlsx)
        return 0

    finally:
        db.close()


if __name__ == "__main__":
    raise SystemExit(main())
